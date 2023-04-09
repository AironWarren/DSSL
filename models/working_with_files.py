from openpyxl import load_workbook
import os
import allure

py_file = os.path.abspath(__file__)
dssl_dir = os.path.dirname(py_file)
resources_dir = os.path.abspath(os.path.join(dssl_dir, '..', 'resources'))


class WorkingWithFiles:
    @allure.step('Counting the number of lines in the xlsx file')
    def number_of_lines_in_the_file(self, way):
        row_in_xlsx = load_workbook(way).active.max_row
        return row_in_xlsx

    @allure.step('Deleting the xlsx file')
    def del_file(self, way):
        os.remove(way)

